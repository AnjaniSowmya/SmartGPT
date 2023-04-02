import pandas as pd
import re
from sklearn.feature_extraction.text import TfidfVectorizer
from sklearn.metrics.pairwise import cosine_similarity
import openpyxl
from string import Template

# Create a TF-IDF vectorizer
vectorizer = TfidfVectorizer()

# Define a function to preprocess text data
def preprocess(text):
    # Remove punctuation and special characters
    text = re.sub('[^A-Za-z0-9]+', ' ', text)
    # Return preprocessed text
    return text

def get_rel_type(relevance):
    type = ''
    # load excel with its path
    wrkbk = openpyxl.load_workbook("gptdata.xlsx")
    sh = wrkbk["RelevanceTypes"]
  
    # iterate through excel and display data
    for i in range(2, sh.max_row+1):
        name = sh.cell(row=i, column=1).value
        #print(name)
        if name == relevance:
            type = sh.cell(row=i, column=2).value
            break
    
    return type

# Define a function to retrieve responses from the sheet based on user input
def get_similarity(relevance_vector, input_text, allmax=False):
    # Preprocess the input text
    input_text = preprocess(input_text)
    # Convert the input text to a TF-IDF vector
    input_vector = vectorizer.transform([input_text])

    # Compute the cosine similarity between the input vector and the document vectors
    similar = True
    # print(input_vector)
    # print(relevance_vector)
    cosine_similarities = cosine_similarity(input_vector, relevance_vector)
    # print(cosine_similarities)
    # print(cosine_similarities.argmax())
    # print(cosine_similarities.any())
    # print(cosine_similarities.all())
    # print(cosine_similarities.argmin())
    if cosine_similarities.any()==0:
        similar = False
    if not similar:
        return -1
    
    # Get the index of the most similar document
    max_index = cosine_similarities.argmax()
    if allmax == True:
        indexes = []
        for i in range(0, len(cosine_similarities[0])):
            if cosine_similarities[0][i] == cosine_similarities[0][max_index]:
                indexes.append(i)
        return indexes
    return max_index

def get_answer(question, rel_name=''):
    # Load data from Excel sheet
    relevance = pd.read_excel('gptdata.xlsx', sheet_name='Relevance')
    # Preprocess data
    relevance['Question'] = relevance['Question'].apply(preprocess)
    # Fit the vectorizer to the data
    relevance_vector = vectorizer.fit_transform(relevance['Question'])
    skip_question = False
    
    # Get response from the sheet based on user input
    relset = False
    if rel_name=='':
        rel_index = get_similarity(relevance_vector, question)
        if rel_index==-1:
            return {'answer': 'Sorry, I could not understand your question. Please rephrase and try again.', 'rel_name': ''}
        rel_name = relevance.iloc[rel_index]['Relevance']
        #print('Relevance:', rel_name)
    else:
        relset = True
    type = get_rel_type(rel_name)
    #print('Type: ', type)
    if type=='QNA':
        # Load data from Excel sheet
        rel_sheet = pd.read_excel('gptdata.xlsx', sheet_name=rel_name)
        # Preprocess data
        rel_sheet['Question'] = rel_sheet['Question'].apply(preprocess)
        # Fit the vectorizer to the data
        rel_sheet_vector = vectorizer.fit_transform(rel_sheet['Question'])
        index = get_similarity(rel_sheet_vector, question)
        if index==-1:
            if relset==True:
                return get_answer(question)
            else:
                return {'answer': 'Sorry, I could not understand your question. Please rephrase and try again.', 'rel_name': ''}
        rel_name = rel_sheet.iloc[index]['Relevance']
        return {'answer': rel_sheet.iloc[index]['Answer'], 'rel_name': rel_name}
    elif type=='DataSheet':
        # Load data from Excel sheet
        rel_sheet = pd.read_excel('gptdata.xlsx', sheet_name='DataSheet')
        # print(rel_sheet)
        rel_data_sheet = rel_sheet[rel_sheet['Sheet']==rel_name]
        # print(rel_data_sheet)
        rel_data_sheet['Question'] = rel_data_sheet['Question'].apply(preprocess)
        # print(rel_data_sheet['Question'])
        # Fit the vectorizer to the data
        rel_data_sheet_vector = vectorizer.fit_transform(rel_data_sheet['Question'])
        index = get_similarity(rel_data_sheet_vector, question)
        if index==-1:
            return {'answer': 'Sorry, I could not understand your question. Please rephrase and try again.', 'rel_name': ''}
        type = rel_data_sheet.iloc[index]['Type']
        output = rel_data_sheet.iloc[index]['Output']
        unique = rel_data_sheet.iloc[index]['Unique']
        #print('Question is of type: ' + type + ' and it is related to: ' + rel_data_sheet.iloc[index]['Name'])
        search_in = rel_data_sheet.iloc[index]['Input'].split(',')
        # print(search_in)
        rel_details_sheet = pd.read_excel('gptdata.xlsx', sheet_name=rel_name)
        combined_details = ''
        for i in range(len(search_in)):
            #print('Input: '+ search_in[i])
            combined_details+=' '+rel_details_sheet[search_in[i]]
        combined_details = combined_details.apply(preprocess)
        combined_details_vector = vectorizer.fit_transform(combined_details)
        dindex = get_similarity(combined_details_vector, question, True)
        #print(combined_details)
        if dindex==-1:
            return {'answer': 'Sorry, I could not understand your question. Please rephrase and try again.', 'rel_name': ''}
        else:
            rds1 = rel_details_sheet.filter(items=dindex, axis=0)
            rds2 = rds1.drop_duplicates(subset = unique)
            t = Template(output)
            values = { 'NL': '\n' }
            for i in range(len(search_in)):
                values[search_in[i]] = rel_details_sheet.iloc[dindex[0]][search_in[i]]
            if type == 'Count':
                values['Count'] = len(rds2)
            elif type == 'List':
                list = ''
                for i in range(len(rds2)):
                    list += rds2.iloc[i][unique] + '\n'
                values['List'] = list
            return {'answer': t.substitute(values), 'rel_name': ''}